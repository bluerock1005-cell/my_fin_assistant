#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""银行全称 → 21家银行白名单分类，输出 Excel。

用法：
    python bank_classify.py                          # 使用内置测试数据
    python bank_classify.py 银行清单.xlsx               # 读取 xlsx/csv/txt，自动识别"银行全称"列
    python bank_classify.py 银行清单.xlsx -c 2          # 指定第2列为银行全称（1-based）
    python bank_classify.py 银行清单.xlsx -o 结果.xlsx   # 指定输出路径

输入文件支持：
    .xlsx / .xlsm  —— 自动查找包含"银行"字样的表头列；找不到则取第1列
    .csv           —— 同上，自动探测 UTF-8 / GBK 编码
    .txt           —— 每行一个银行全称
未提供输入文件时，使用脚本内置的 TEST_BANKS 作为示例数据。
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===== 21家银行白名单：[检测子串, 简称] =====
# 注意特殊映射：邮储银行=邮政储蓄；浦发银行=含"浦发"/"浦东发展"
# 按子串长度从长到短排列，避免短子串抢先误命中更具体的名称。
WHITELIST = sorted(
    [
        ("光大银行", "光大银行"),
        ("平安银行", "平安银行"),
        ("华夏银行", "华夏银行"),
        ("中信银行", "中信银行"),
        ("邮政储蓄", "邮储银行"),
        ("交通银行", "交通银行"),
        ("招商银行", "招商银行"),
        ("兴业银行", "兴业银行"),
        ("工商银行", "工商银行"),
        ("中国银行", "中国银行"),
        ("建设银行", "建设银行"),
        ("农业银行", "农业银行"),
        ("民生银行", "民生银行"),
        ("浦发", "浦发银行"),  # "上海浦发发展银行" 等
        ("浦东发展", "浦发银行"),  # "上海浦东发展银行"官方全称；"浦发"子串无法命中"浦东发展"，必须单独补一条
        ("浙商银行", "浙商银行"),
        ("宁波银行", "宁波银行"),  # "宁波通商银行" 不含"宁波银行"，不会误命中
        ("江苏银行", "江苏银行"),  # "江苏苏商银行" 不含"江苏银行"，不会误命中
        ("广发银行", "广发银行"),
        ("上海银行", "上海银行"),
        ("南京银行", "南京银行"),
        ("北京银行", "北京银行"),
    ],
    key=lambda x: len(x[0]),
    reverse=True,
)

# ---------------------------------------------------------------------------
# 21家银行白名单展示（用于客户端 UI 展示，按类型分组）
# 与上方 WHITELIST 同源：展示列表的去重简称必须等于 WHITELIST 实际命中的简称集合，
# 二者由 validate_whitelist() 校验一致性，避免两处不同步。
# ---------------------------------------------------------------------------
WHITELIST_GROUPS = [
    ("国有大型商业银行", [
        "工商银行", "农业银行", "中国银行", "建设银行", "交通银行", "邮储银行",
    ]),
    ("股份制商业银行", [
        "中信银行", "光大银行", "华夏银行", "民生银行", "招商银行",
        "兴业银行", "平安银行", "浦发银行", "广发银行", "浙商银行",
    ]),
    ("城市商业银行", [
        "北京银行", "南京银行", "宁波银行", "江苏银行", "上海银行",
    ]),
]


def get_whitelist_groups() -> list[tuple[str, list[str]]]:
    """返回 [(类型, [银行简称...]), ...]，供 UI 分组展示。"""
    return WHITELIST_GROUPS


def get_whitelist_banks() -> list[str]:
    """返回扁平化的 21 家银行简称列表（保持展示顺序）。"""
    return [bank for _, banks in WHITELIST_GROUPS for bank in banks]


def validate_whitelist() -> bool:
    """校验展示列表与 WHITELIST 实际命中的银行一致。

    若有人改动 WHITELIST 的模式/简称却忘了同步 WHITELIST_GROUPS，
    会抛出 AssertionError，便于在测试/启动时发现问题。
    """
    display = set(get_whitelist_banks())
    actual = {short for _, short in WHITELIST}
    if display != actual:
        missing = sorted(actual - display)
        extra = sorted(display - actual)
        raise AssertionError(
            f"白名单展示列表与 WHITELIST 不一致：缺失={missing} 多余={extra}"
        )
    return True


FONT_NAME = "微软雅黑"

# 仅在未提供输入文件时使用的示例数据
TEST_BANKS = [
    "宁波通商银行股份有限公司杭州分行",
    "中国民生银行股份有限公司杭州庆春支行",
]


def classify(name: str) -> tuple[str, str]:
    """返回 (所属21家银行简称 或 '', 分类标签)。"""
    for pat, short in WHITELIST:
        if pat in name:
            return short, "21银行承兑汇票"
    return "", "非21银行承兑汇票"


def _read_text_any_encoding(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法识别文件编码: {path}")


def _pick_column_index(header: list, col_arg: int | None) -> int:
    if col_arg is not None:
        return col_arg - 1
    for i, h in enumerate(header):
        if h and "银行" in str(h):
            return i
    return 0


def _load_from_xlsx(path: Path, col_arg: int | None) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    if not rows:
        return []
    idx = _pick_column_index(list(rows[0]), col_arg)
    # 若首行看起来是表头（该列不含"银行"关键字之外的合理全称长度），跳过首行
    header_val = str(rows[0][idx]) if idx < len(rows[0]) and rows[0][idx] else ""
    data_rows = rows[1:] if ("银行" in header_val and len(header_val) < 10) else rows
    return [str(r[idx]).strip() for r in data_rows if idx < len(r) and r[idx] not in (None, "")]


def _load_from_csv(path: Path, col_arg: int | None) -> list[str]:
    text = _read_text_any_encoding(path)
    reader = list(csv.reader(text.splitlines()))
    reader = [r for r in reader if any(c.strip() for c in r)]
    if not reader:
        return []
    idx = _pick_column_index(reader[0], col_arg)
    header_val = reader[0][idx] if idx < len(reader[0]) else ""
    data_rows = reader[1:] if ("银行" in header_val and len(header_val) < 10) else reader
    return [r[idx].strip() for r in data_rows if idx < len(r) and r[idx].strip()]


def _load_from_txt(path: Path) -> list[str]:
    text = _read_text_any_encoding(path)
    return [line.strip() for line in text.splitlines() if line.strip()]


def _read_clipboard() -> str:
    # 优先使用 tkinter（标准库），若不可用再尝试 pyperclip，最后在 Windows 下尝试 PowerShell
    try:
        import tkinter as tk
    except Exception:
        tk = None

    if tk is not None:
        try:
            root = tk.Tk()
            root.withdraw()
            try:
                return root.clipboard_get()
            finally:
                root.destroy()
        except Exception:
            # 若 tkinter 存在但读取失败，继续尝试其他方案
            pass

    # 尝试 pyperclip（第三方库），若已安装则使用
    try:
        import pyperclip

        try:
            return pyperclip.paste()
        except Exception:
            pass
    except Exception:
        pass

    # Windows 平台：尝试通过 PowerShell 读取剪贴板（不依赖额外 Python 包）
    try:
        import subprocess
        import sys

        if sys.platform.startswith("win"):
            proc = subprocess.run(["powershell", "-NoProfile", "-Command", "Get-Clipboard"], capture_output=True, text=True)
            if proc.returncode == 0:
                return proc.stdout
    except Exception:
        pass

    raise RuntimeError("无法读取剪贴板：请安装 tkinter 或 pyperclip，或在 Windows 上确保 PowerShell 可用以读取剪贴板")


def load_banks(input_path: str | None, col_arg: int | None, paste: bool = False) -> list[str]:
    if paste:
        text = _read_clipboard()
        return [line.strip() for line in text.splitlines() if line.strip()]
    if input_path is None:
        return TEST_BANKS
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"输入文件不存在: {path}")
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _load_from_xlsx(path, col_arg)
    if ext == ".csv":
        return _load_from_csv(path, col_arg)
    return _load_from_txt(path)


def build_workbook(banks: list[str], out_path: Path) -> tuple[int, int]:
    rows = []
    for i, name in enumerate(banks, 1):
        short, cat = classify(name)
        rows.append((i, name, short, cat))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "银行承兑汇票分类"

    base_font = Font(name=FONT_NAME)
    ws.sheet_view.showGridLines = False

    # 标题
    ws.merge_cells("A1:D1")
    ws["A1"] = "银行承兑汇票分类结果（21家银行白名单）"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # 表头
    headers = ["序号", "银行全称", "所属21家银行", "分类"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    thin_hdr = Side(style="thin", color="BFBFBF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(thin_hdr, thin_hdr, thin_hdr, thin_hdr)

    # 数据行：预注册 NamedStyle 并按引用套用，避免每个单元格都重新构造
    # Font/Alignment/Border 对象触发 openpyxl 内部样式去重的重复哈希开销
    # （5万行规模下这一步能把耗时从约50s降到几秒）。
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _reg(name, font, alignment, fill=None):
        style = NamedStyle(
            name=name, font=font, alignment=alignment, border=border, fill=fill or PatternFill()
        )
        if name not in wb.named_styles:
            wb.add_named_style(style)
        return name

    style_left = _reg("bc_left", base_font, Alignment(vertical="center", horizontal="left"))
    style_center = _reg("bc_center", base_font, Alignment(vertical="center", horizontal="center"))
    style_cat_yes = _reg(
        "bc_cat_yes",
        Font(name=FONT_NAME, bold=True, color="375623"),
        Alignment(vertical="center", horizontal="center"),
        PatternFill("solid", fgColor="E2EFDA"),
    )
    style_cat_no = _reg(
        "bc_cat_no",
        Font(name=FONT_NAME, bold=True, color="843C0C"),
        Alignment(vertical="center", horizontal="center"),
        PatternFill("solid", fgColor="FCE4D6"),
    )

    for r, (idx, name, short, cat) in enumerate(rows, start=3):
        row_styles = (
            style_center,
            style_left,
            style_center,
            style_cat_yes if cat == "21银行承兑汇票" else style_cat_no,
        )
        for c, (v, sty) in enumerate(zip((idx, name, short, cat), row_styles), 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.style = sty

    last_data_row = 2 + len(rows) if rows else 2

    # 冻结表头 + 筛选
    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = f"A2:D{last_data_row}"

    # 列宽：按内容自适应（含中文宽字符估算）
    def _display_width(s: str) -> float:
        return sum(1.8 if ord(ch) > 0x2E80 else 1.0 for ch in s)

    widths = {1: len("序号"), 2: len("银行全称"), 3: len("所属21家银行"), 4: len("分类")}
    for idx, name, short, cat in rows:
        widths[1] = max(widths[1], len(str(idx)))
        widths[2] = max(widths[2], int(_display_width(name)))
        widths[3] = max(widths[3], int(_display_width(short)))
        widths[4] = max(widths[4], int(_display_width(cat)))
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 4, 8), 60)

    # 统计表（使用公式，随数据自动更新）
    srow = last_data_row + 3
    ws.cell(row=srow, column=1, value="统计").font = Font(name=FONT_NAME, bold=True, size=12)
    stat_labels = ["总条数", "21银行承兑汇票", "非21银行承兑汇票"]
    if rows:
        data_range = f"D3:D{last_data_row}"
        stat_formulas = [
            f"=COUNTA(B3:B{last_data_row})",
            f'=COUNTIF({data_range},"21银行承兑汇票")',
            f'=COUNTIF({data_range},"非21银行承兑汇票")',
        ]
    else:
        stat_formulas = [0, 0, 0]
    for i, (label, val) in enumerate(zip(stat_labels, stat_formulas), start=srow + 1):
        ws.cell(row=i, column=1, value=label).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=i, column=2, value=val).font = base_font

    # 非21名单汇总
    nrow = srow + len(stat_labels) + 2
    ws.cell(row=nrow, column=1, value="非21家银行明细").font = Font(
        name=FONT_NAME, bold=True, size=12, color="843C0C"
    )
    rr = nrow + 1
    for idx, name, short, cat in rows:
        if cat == "非21银行承兑汇票":
            ws.cell(row=rr, column=1, value=name).font = base_font
            rr += 1

    wb.save(out_path)

    n_yes = sum(1 for *_, cat in rows if cat == "21银行承兑汇票")
    n_no = len(rows) - n_yes
    return n_yes, n_no


if __name__ == "__main__":
    import argparse
    import sys
    from pathlib import Path as _Path

    _parser = argparse.ArgumentParser(description="银行全称 → 21家银行白名单分类工具")
    _parser.add_argument(
        "input", nargs="?", default=None, help="输入文件路径 (.xlsx/.csv/.txt)，留空则使用内置示例数据"
    )
    _parser.add_argument("-c", "--column", type=int, default=None, help="银行全称所在列号（从1开始）")
    _parser.add_argument("-o", "--output", default=None, help="输出文件路径，默认为当前目录下 银行承兑汇票分类.xlsx")
    _parser.add_argument("--paste", action="store_true", help="从系统剪贴板读取银行名称（每行一条），优先于输入文件")
    _args = _parser.parse_args()

    if _args.paste and _args.input is not None:
        print("参数错误：不能同时指定输入文件和 --paste")
        sys.exit(1)

    try:
        _banks = load_banks(_args.input, _args.column, _args.paste)
    except Exception as _e:
        print(f"读取输入失败: {_e}")
        sys.exit(1)

    if not _banks:
        print("未读取到任何银行名称，请检查输入文件内容或使用 -c 指定正确的列号。")
        sys.exit(1)

    _out = _Path(_args.output) if _args.output else _Path.cwd() / "银行承兑汇票分类.xlsx"
    _n_yes, _n_no = build_workbook(_banks, _out)

    print(f"已生成: {_out}")
    print(f"总计 {len(_banks)} 条 | 21银行承兑汇票 {_n_yes} 条 | 非21银行承兑汇票 {_n_no} 条")
