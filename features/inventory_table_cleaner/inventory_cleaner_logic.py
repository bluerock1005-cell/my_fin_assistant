#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""features/inventory_table_cleaner/inventory_cleaner_logic.py — 存货收发存汇总表清洗逻辑层。

实现方式：通过 win32com 驱动本机 **Microsoft Excel** 完成表头清洗（COM 自动化）。

处理的原始表头为 3 层合并结构（测试文档 `存货收发存汇总表测试.xlsx` 的 `zy` 表）：
  - 第 3 行：分组名（期初结存 / 本期收入 / 本期发出 / 本期结存），左侧 A~K 为竖向合并的单值列
  - 第 4 行：中间层（期初结存/本期结存的 数量/单价/金额；本期收入/本期发出的子类如采购类/生产类…）
  - 第 5 行：本期收入/本期发出的 数量/单价/金额

清洗步骤（与需求一致）：
  1) 删除第 1、2 行（标题/本位币说明）
  2) 取消第 3~5 行合并单元格（先按合并区域把标签填充到每个单元格，再 UnMerge，保留层级信息）
  3) 把分组名与 数量/单价/金额 等逐层结合成完整字段名（如期初结存-数量、本期收入-采购类-数量）
  4) 删除多余的中间/底行，得到单层（或双层）干净表头
  5) 删除数据区的「合计/总计」汇总行（通常为数据区最后一行）
  6) 若指定「保留期初/期末会计期间」，清空其余期间对应的期初结存/期末结存数据单元格
     （数量/单价/金额），**保留列不动**——只清数据、不删列

为什么用 COM 而非 openpyxl：COM 由 Excel 自身执行删除/取消合并/保存，**完整保留原有
数字格式、列宽、字体等样式**，且不触发 openpyxl 重写带来的样式丢失。

UI 与逻辑严格分离：本文件不依赖任何 Qt 类型；`win32com` 为惰性导入，
缺少 Excel / COM 时模块仍可 import（仅运行时会抛出清晰错误）。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

# ===== COM 常量（用字面量，避免依赖 gencache 生成的枚举常量） =====
_XL_FILE_FORMAT_XLSX = 51  # xlOpenXMLWorkbook：.xlsx


def _com_excel_available() -> bool:
    """本机是否具备 COM Excel 自动化所需的 pywin32。"""
    try:
        import win32com  # noqa: F401
        return True
    except Exception:
        return False


def _create_excel_app(visible: bool = False):
    """创建并初始化 Excel.Application 实例。

    必须在调用线程内调用（内部负责 CoInitialize）；返回的应用实例
    由调用方在用完后通过 :func:`_quit_excel_app` 释放。

    Raises:
        RuntimeError: 无法启动 Excel（未安装 / pywin32 缺失 / 非 Windows）。
    """
    import pythoncom
    import win32com.client

    # 同一线程重复 CoInitialize 会抛错，忽略即可
    try:
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
    except pythoncom.error:
        pass

    try:
        # DispatchEx 强制创建独立的新实例，避免复用用户已打开的 Excel
        # 或上一轮刚退出、尚在过渡态的残留实例（复用会导致后续属性设置失败）。
        app = win32com.client.DispatchEx("Excel.Application")
    except Exception as exc:  # noqa: BLE001
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        raise RuntimeError(
            "无法启动 Microsoft Excel（COM 自动化失败）。请确认：\n"
            "  1) 当前为 Windows 且已安装 Microsoft Excel；\n"
            "  2) 已安装 pywin32（pip install pywin32）；\n"
            "  3) Excel 未被其他进程独占锁定。"
        ) from exc

    # 以下属性设置个别环境可能抛 COM 异常（如共享实例状态异常），忽略即可，
    # 不影响后台静默处理。
    try:
        app.Visible = visible
    except Exception:
        pass
    try:
        app.DisplayAlerts = False   # 抑制保存/覆盖等弹窗
    except Exception:
        pass
    try:
        app.ScreenUpdating = False  # 后台静默处理，提速
    except Exception:
        pass
    return app


def _quit_excel_app(app) -> None:
    """退出 Excel 并释放当前线程的 COM 初始化。对 None 安全。"""
    if app is None:
        return
    try:
        app.Quit()
    except Exception:
        pass
    try:
        import pythoncom
        pythoncom.CoUninitialize()
    except Exception:
        pass


def _as_text(value) -> str:
    """把 COM 单元格值规整为字符串（None / 数字 / 字符串统一处理）。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_accounting_periods(input_path: str | Path) -> list[str]:
    """读取源 Excel B 列（会计期间），去重 + 排序后返回期间字符串列表。

    返回值示例：``["01", "02", "03", "04", "05", "06"]``。

    内部值保留原始的零填充字符串（如 ``"01"``），方便与清洗后表头中嵌入的期号
    （如「期初结存-01-数量」里的 ``"01"``）直接做字符串匹配。

    处理要点：
      - 只取「纯数字串」单元格（如 ``"01"``），自动跳过标题行 / 表头标签行
        （这些单元格是长文本或「会计期间」字样，不匹配纯数字模式）；
      - 兼容合并单元格：数据区 B 列可能被标题行的合并区域覆盖，通过合并区域
        锚点值回填，确保合并单元格也能读到正确期间；
      - 归一化为 2 位零填充字符串（``"1"`` -> ``"01"``，``"01"`` -> ``"01"``），
        保证与下游表头期号一致。
    """
    from openpyxl import load_workbook

    p = Path(input_path).resolve()
    if not p.exists():
        raise FileNotFoundError(f"源文件不存在：{p}")

    wb = load_workbook(p, data_only=True)
    try:
        ws = wb.active
        # 建立合并单元格锚点值映射（仅关心 B 列，列号 2）
        anchor: dict[tuple[int, int], str] = {}
        for rng in list(ws.merged_cells.ranges):
            if not (rng.min_col <= 2 <= rng.max_col):
                continue
            val = ws.cell(rng.min_row, rng.min_col).value
            sval = "" if val is None else str(val).strip()
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    anchor[(r, c)] = sval

        seen: set[str] = set()
        out: list[str] = []
        for r in range(1, ws.max_row + 1):
            if (r, 2) in anchor:
                s = anchor[(r, 2)]
            else:
                v = ws.cell(r, 2).value
                s = "" if v is None else str(v).strip()
            # 仅保留纯数字串（会计期间，如 "01"）；跳过标题 / 表头标签
            if not s or not re.fullmatch(r"\d{1,3}", s):
                continue
            norm = f"{int(s):02d}"
            if norm not in seen:
                seen.add(norm)
                out.append(norm)
        out.sort(key=lambda x: int(x))
        return out
    finally:
        wb.close()


def _find_col_index(headers: list[str], target: str) -> int | None:
    """在扁平表头中查找目标字段，返回 1-based Excel 列号；未找到返回 None。"""
    for i, h in enumerate(headers):
        if h == target:
            return i + 1
    return None


def _find_period_cols(headers: list[str], base: str, sep: str) -> list[int] | None:
    """查找 base-数量/单价/金额 三列，返回 1-based 列号列表；缺任一列返回 None。"""
    cols: list[int] = []
    for suffix in ("数量", "单价", "金额"):
        idx = _find_col_index(headers, f"{base}{sep}{suffix}")
        if idx is None:
            return None
        cols.append(idx)
    return cols


def _find_open_cols(headers: list[str], sep: str) -> list[int] | None:
    """查找期初结存-数量/单价/金额三列。"""
    return _find_period_cols(headers, "期初结存", sep)


def _find_close_cols(headers: list[str], sep: str) -> list[int] | None:
    """查找期末结存/本期结存-数量/单价/金额三列（优先期末结存，其次本期结存）。"""
    for base in ("期末结存", "本期结存"):
        cols = _find_period_cols(headers, base, sep)
        if cols is not None:
            return cols
    return None


def _resolve_close_cols(headers: list[str], sep: str) -> list[int] | None:
    """本期结存（期末/本期结存）列：优先固定 BC(55)/BD(56)/BE(57)，回退动态查找。

    用户确认的源表「本期结存-数量/单价/金额」位于 BC/BD/BE 三列。显式锁定这些
    固定列可避免「多期间、表头带期号（如期末结存-06-数量）」时动态表头匹配失效、
    而静默跳过期末清空的问题。

    仅当固定范围超出表宽（表头数不足 57 列）或 BC/BD/BE 表头不含「结存」字样时，
    才回退到 :func:`_find_close_cols` 按表头查找，保证鲁棒性。
    """
    fixed = [55, 56, 57]  # BC, BD, BE（1-based 列号）
    if len(headers) >= max(fixed):
        ok = True
        for c in fixed:
            h = headers[c - 1] if 0 <= c - 1 < len(headers) else ""
            if h and "结存" not in h:
                ok = False
                break
        if ok:
            return fixed
    return _find_close_cols(headers, sep)


def _clear_non_kept_period_rows(
    ws,
    headers: list[str],
    header_rows: int,
    keep_open: str | None,
    keep_close: str | None,
    sep: str,
    log,
) -> int:
    """按行清空非保留期间的期初/期末结存数据单元格（列保留不动）。

    与需求一致的两步：
      1. 定位要清空的行：遍历数据区，读取该行的「会计期间」列；
         若等于保留期间则跳过，否则该行即为目标行。
      2. 清空目标行的余额单元格：把该行「期初结存-数量/单价/金额」
         （以及「期末结存/本期结存-数量/单价/金额」）三格清空。
    本期收入 / 本期发出 / 固定列完全不动。

    keep_open/keep_close 为归一化的 2 位零填充字符串（如 "01"），None 表示不筛选。
    返回被清空的数据单元格总数。
    """
    period_col = _find_col_index(headers, "会计期间")
    if period_col is None:
        log("⚠ 未找到「会计期间」列，无法按行清空期间数据")
        return 0

    open_cols = _find_open_cols(headers, sep)
    # 结存（期末/本期结存）列：优先固定 BC/BD/BE（用户确认的位置），回退动态查找
    close_cols = _resolve_close_cols(headers, sep)

    if keep_open is None and keep_close is None:
        return 0
    if keep_open is not None and open_cols is None:
        log("⚠ 未找到「期初结存」数量/单价/金额列，跳过期初清空")
    if keep_close is not None and close_cols is None:
        log("⚠ 未找到「期末结存/本期结存」数量/单价/金额列（BC/BD/BE），跳过期末清空")

    first_data_row = header_rows + 1
    last_row = int(ws.UsedRange.Rows.Count)
    if last_row < first_data_row:
        return 0

    cleared = 0
    for r in range(first_data_row, last_row + 1):
        val = ws.Cells(r, period_col).Value
        if val is None:
            continue
        period_val = str(val).strip()
        if not period_val.isdigit():
            continue
        period_val = f"{int(period_val):02d}"

        if keep_open is not None and open_cols is not None and period_val != keep_open:
            for c in open_cols:
                ws.Cells(r, c).ClearContents()
                cleared += 1
        if keep_close is not None and close_cols is not None and period_val != keep_close:
            for c in close_cols:
                ws.Cells(r, c).ClearContents()
                cleared += 1

    if cleared:
        log(f"🧹 已清空 {cleared} 个非保留期间余额单元格（按行判断，列保留不动）")
    return cleared


@dataclass
class CleanResult:
    """单次清洗结果。"""
    input_path: Path
    output_path: Path
    header_mode: str
    header_count: int
    headers: list[str]
    data_rows: int
    ok: bool
    error: str | None = None
    log_lines: list[str] = field(default_factory=list)
    deleted_columns: int = 0
    deleted_rows: int = 0
    cleared_cells: int = 0


def process_inventory(
    input_path: str | Path,
    output_path: str | Path,
    header_mode: str = "single",
    sep: str = "-",
    progress_callback: Callable[[str], None] | None = None,
    keep_open_period: str | None = None,
    keep_close_period: str | None = None,
) -> CleanResult:
    """清洗存货收发存汇总表，生成单层（或双层）表头的新 Excel。

    Args:
        input_path: 源 .xlsx 路径。
        output_path: 目标 .xlsx 路径（父目录会被创建；会另存，不覆盖源文件）。
        header_mode: "single" 单层表头（合并成一行，删除原第 4、5 行）；
                     "double" 双层表头（保留分组行 + 合并后的明细行，仅删原第 5 行）。
        sep: 层级结合的分隔符，默认 "-" 。
        progress_callback: 进度日志回调，用于向 UI 回传消息。
        keep_open_period: 保留的「期初会计期间」（如 "01"，零填充字符串）。非 None 时按行读取
            「会计期间」列；不等于该期间的行，其「期初结存-数量/单价/金额」单元格被清空，
            **列保留不动**；仅保留该期余额数据。
        keep_close_period: 保留的「期末会计期间」（如 "06"，零填充字符串）。非 None 时按行读取
            「会计期间」列；不等于该期间的行，其「期末结存/本期结存-数量/单价/金额」单元格
            被清空，**列保留不动**。本期收入 / 本期发出等流转列完全不动。

    Returns:
        CleanResult 汇总对象。

    Note:
        调用线程需已 CoInitialize（由 :func:`_create_excel_app` 处理）。
    """
    def _log(msg: str) -> None:
        if progress_callback is not None:
            progress_callback(msg)

    if not _com_excel_available():
        raise RuntimeError(
            "本机缺少 pywin32，无法进行 Excel COM 处理。请执行：pip install pywin32"
        )

    in_p = Path(input_path).resolve()
    out_p = Path(output_path).resolve()
    if not in_p.exists():
        raise FileNotFoundError(f"源文件不存在：{in_p}")
    if in_p == out_p:
        raise ValueError("输出文件不能与源文件相同（避免覆盖原文件）")
    out_p.parent.mkdir(parents=True, exist_ok=True)
    if out_p.suffix.lower() != ".xlsx":
        out_p = out_p.with_suffix(".xlsx")

    header_mode = "single" if header_mode != "double" else "double"

    app = _create_excel_app(False)
    try:
        _log(f"打开源文件：{in_p.name}")
        wb = app.Workbooks.Open(str(in_p))
        try:
            # 处理第一个工作表（测试文档为 zy 表）
            ws = wb.Worksheets(1)

            # 最大列（合并/数据都不会超过此范围）
            maxcol = int(ws.UsedRange.Columns.Count)
            if maxcol <= 0:
                maxcol = 1
            maxcol = min(maxcol, 1000)

            # 步骤 1：删除第 1、2 行（上移）
            _log("步骤1：删除第 1、2 行（标题/说明）")
            ws.Rows("1:2").Delete()

            # 此时原表头 3~5 行已上移到 1~3 行
            # 步骤 2：先按合并区域把标签填充到每个单元格（保留层级），再取消合并
            _log("步骤2：读取合并区域并填充层级标签，随后取消第 3~5 行合并")
            filled: dict[tuple[int, int], str] = {}
            for r in (1, 2, 3):
                for c in range(1, maxcol + 1):
                    cell = ws.Cells(r, c)
                    if cell.MergeCells:
                        ma = cell.MergeArea
                        # 只在合并区左上角处理一次，避免重复
                        if ma.Row == r and ma.Column == c:
                            val = _as_text(cell.Value)
                            for rr in range(ma.Row, ma.Row + ma.Rows.Count):
                                for cc in range(ma.Column,
                                                ma.Column + ma.Columns.Count):
                                    if 1 <= rr <= 3 and 1 <= cc <= maxcol:
                                        filled[(rr, cc)] = val
                    else:
                        filled[(r, c)] = _as_text(cell.Value)

            # 取消合并（仅作用于当前 1~3 行区域）
            ws.Range(ws.Cells(1, 1), ws.Cells(3, maxcol)).UnMerge()

            # 把填充后的标签写回每个单元格（取消合并后这些原本空白的单元格需显式赋值）
            for (r, c), v in filled.items():
                ws.Cells(r, c).Value = v

            # 步骤 3：逐列把 1~3 行层级结合成完整字段名
            _log("步骤3：结合分组名与 数量/单价/金额 等，生成完整字段名")
            combined: list[str] = []
            for c in range(1, maxcol + 1):
                parts: list[str] = []
                for r in (1, 2, 3):
                    v = filled.get((r, c), "")
                    if v:
                        parts.append(v)
                # 去除层级冗余：
                #  - 连续完全相同则去重；
                #  - 子级名已包含父级名（如 本期收入 / 本期收入合计）则只保留更具体的，
                #    避免 本期收入-本期收入合计-数量 这类冗余，得到 本期收入合计-数量。
                deduped: list[str] = []
                for p in parts:
                    if not deduped:
                        deduped.append(p)
                    else:
                        last = deduped[-1]
                        if p == last:
                            continue
                        elif p.startswith(last):
                            deduped[-1] = p
                        elif last.startswith(p):
                            pass
                        else:
                            deduped.append(p)
                combined.append(sep.join(deduped) if deduped else "")

            # 步骤 4：写回表头并删除多余行
            if header_mode == "single":
                _log("步骤4：写回单层表头，并删除原第 4、5 行")
                for c in range(1, maxcol + 1):
                    ws.Cells(1, c).Value = combined[c - 1]
                ws.Rows("2:3").Delete()
                header_rows = 1
            else:
                _log("步骤4：写回明细表头（第 2 行），保留分组行（第 1 行），删除原第 5 行")
                for c in range(1, maxcol + 1):
                    ws.Cells(2, c).Value = combined[c - 1]
                ws.Rows(3).Delete()
                header_rows = 2

            # 步骤 4b：删除数据区的「合计/总计」汇总行（通常为数据区最后一行）
            # 识别方式：扫描数据区第 1 列，凡含「合计」/「总计」的行即删除（从下往上删避免偏移）。
            _log("步骤4b：删除数据区合计/总计行")
            deleted_rows = 0
            data_start = header_rows + 1
            last_data_row = int(ws.UsedRange.Rows.Count)
            for r in range(last_data_row, data_start - 1, -1):
                a_val = _as_text(ws.Cells(r, 1).Value)
                if a_val and ("合计" in a_val or "总计" in a_val):
                    ws.Rows(r).Delete()
                    deleted_rows += 1
                    _log(f"  已删除合计行（原第 {r} 行）")
            if deleted_rows:
                _log(f"✂ 共删除 {deleted_rows} 行合计/总计行")

            # === 步骤5：按保留的会计期间，按行清空非保留期间余额数据单元格（列保留不动）===
            cleared_count = 0
            if keep_open_period is not None or keep_close_period is not None:
                cleared_count = _clear_non_kept_period_rows(
                    ws, combined, header_rows,
                    keep_open_period, keep_close_period, sep, _log)

            # 重新统计最终表头与数据行数（清空数据不改变列/行结构，仍重新读取以确保准确）
            final_maxcol = int(ws.UsedRange.Columns.Count)
            if header_mode == "single":
                headers = [_as_text(ws.Cells(1, c).Value)
                           for c in range(1, final_maxcol + 1)]
            else:
                # 双层：以明细行（第 2 行）作为预览表头
                headers = [_as_text(ws.Cells(2, c).Value)
                           for c in range(1, final_maxcol + 1)]
            total_rows = int(ws.UsedRange.Rows.Count)
            data_rows = max(0, total_rows - header_rows)

            # 保存（另存为 .xlsx，不改动源文件）
            _log(f"保存结果：{out_p.name}")
            wb.SaveAs(str(out_p), FileFormat=_XL_FILE_FORMAT_XLSX)

            summary = f"✅ 完成：共 {final_maxcol} 列表头，{data_rows} 行数据"
            extra = []
            if deleted_rows:
                extra.append(f"删除 {deleted_rows} 行合计/总计")
            if cleared_count:
                extra.append(f"清空 {cleared_count} 个非保留期间余额单元格（按行，列保留）")
            summary += ("；" + "、".join(extra) + "。") if extra else "。"
            _log(summary)
            return CleanResult(
                input_path=in_p,
                output_path=out_p,
                header_mode=header_mode,
                header_count=final_maxcol,
                headers=headers,
                data_rows=data_rows,
                ok=True,
                deleted_columns=0,
                deleted_rows=deleted_rows,
                cleared_cells=cleared_count,
                log_lines=[],
            )
        finally:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
    finally:
        _quit_excel_app(app)
