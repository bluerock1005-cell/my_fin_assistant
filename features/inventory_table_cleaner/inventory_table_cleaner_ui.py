#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""features/inventory_table_cleaner/inventory_table_cleaner_ui.py — 存货收发存汇总表清洗界面。

UI 与逻辑严格分离：本文件只管界面与交互，表头清洗放在 inventory_cleaner_logic.py 中。

界面约定参考 features/notes_receivable_import（Fluent 风格）：
  - 拖拽 / 选择源 Excel（.xlsx）
  - 标题 + 操作按钮同一行（左标题，右按钮），按钮统一 objectName="primary"
  - 卡片用 objectName="card"，标题用 pageTitle / cardTitle，说明用 pageDesc
  - 运行日志面板（QPlainTextEdit，objectName="logView"）
  - 后台处理用 ThreadPoolExecutor(max_workers=1) + QTimer 轮询（规避无头 QThread 段错误）
  - 结果预览用自定义只读 QTableView 子类，支持 Ctrl+C 复制
"""
from __future__ import annotations

import concurrent.futures
from datetime import datetime
from pathlib import Path
from queue import Queue

import openpyxl
from PySide6.QtCore import Qt, QTimer, QUrl
from PySide6.QtGui import (
    QDesktopServices,
    QDragEnterEvent,
    QDropEvent,
    QStandardItem,
    QStandardItemModel,
)
from PySide6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLayout,
    QLineEdit,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from core import app_config, theme, utils
from core.feature_base import FeatureModule

from . import inventory_cleaner_logic as _logic


class _ResultTableView(QTableView):
    """结果预览表（只读、可复制）。

    - 双击不可编辑；
    - 选中单元格后 Ctrl+C 复制为 TSV（按选中行列对齐，含表头）；
    - 交替行色，便于核对。
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)
        self.setSelectionBehavior(QTableView.SelectionBehavior.SelectItems)
        self.setSelectionMode(QTableView.SelectionMode.ContiguousSelection)
        self.setAlternatingRowColors(True)
        self.horizontalHeader().setStretchLastSection(False)
        self.setTextElideMode(Qt.TextElideMode.ElideNone)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        if (event.modifiers() & Qt.KeyboardModifier.ControlModifier
                and event.key() == Qt.Key.Key_C):
            self._copy_selection()
        else:
            super().keyPressEvent(event)

    def _copy_selection(self) -> None:
        idxs = self.selectedIndexes()
        if not idxs:
            return
        model = self.model()
        rows = sorted({i.row() for i in idxs})
        cols = sorted({i.column() for i in idxs})
        row_map = {r: i for i, r in enumerate(rows)}
        col_map = {c: i for i, c in enumerate(cols)}
        grid: dict[tuple[int, int], str] = {}
        for i in idxs:
            val = model.data(i, Qt.ItemDataRole.DisplayRole)
            grid[(row_map[i.row()], col_map[i.column()])] = "" if val is None else str(val)
        # 表头行
        header = []
        for c in range(len(cols)):
            hi = model.headerData(cols[c], Qt.Orientation.Horizontal,
                                  Qt.ItemDataRole.DisplayRole)
            header.append("" if hi is None else str(hi))
        lines = ["\t".join(header)]
        for r in range(len(rows)):
            lines.append("\t".join(grid.get((r, c), "") for c in range(len(cols))))
        QApplication.clipboard().setText("\n".join(lines))


class InventoryTableCleanerWidget(QWidget):
    """存货收发存汇总表清洗功能的具体 UI 控件。

    工作流：
      1. 拖拽 / 选择源 Excel（.xlsx）
      2. 选择表头模式（单层/双层）与层级分隔符
      3. 点击「开始处理」→ 后台通过 Excel COM 清洗表头
      4. 在日志面板查看进度，并在结果预览表中核对（前 200 行）
    """

    _PREVIEW_MAX_ROWS = 200

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("page")

        # 状态
        self._input_file: Path | None = None
        self._output_file: Path | None = None

        # 后台线程
        self._executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
        self._future = None
        self._poll_timer = None
        self._log_queue: Queue[str] = Queue()

        self.setAcceptDrops(True)
        self._setup_ui()

    # ====== 布局 ======

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(theme.PAGE_PAD, theme.PAGE_PAD, theme.PAGE_PAD, theme.PAGE_PAD)
        root.setSpacing(theme.CARD_GAP)
        root.setSizeConstraint(QLayout.SetMinimumSize)

        # --- 页头 ---
        title = QLabel("存货收发存汇总表清洗", self)
        title.setObjectName("pageTitle")
        desc = QLabel(
            "通过本机 Microsoft Excel（COM 自动化）清洗存货收发存汇总表的 3 层合并表头："
            "删除标题行、取消合并、把分组名与「数量/单价/金额」结合成完整字段名，"
            "输出干净的单层表头。原文件不会被修改，结果另存为新文件。"
            "需 Windows 且已安装 Excel。",
            self,
        )
        desc.setObjectName("pageDesc")
        desc.setWordWrap(True)
        root.addWidget(title)
        root.addWidget(desc)
        root.addSpacing(theme.SPACING[4])

        # === 卡片1：源文件 + 参数 + 操作 ===
        card1 = QFrame(self)
        card1.setObjectName("card")
        c1 = QVBoxLayout(card1)
        c1.setContentsMargins(theme.SPACING[16], theme.SPACING[16],
                              theme.SPACING[16], theme.SPACING[16])
        c1.setSpacing(theme.SPACING[12])

        # 标题行：左「源文件（拖拽 / 选择）」，右 操作按钮
        h_top = QHBoxLayout()
        h_top.setSpacing(theme.SPACING[12])
        h_lbl = QLabel("源文件（拖拽 / 选择）", card1)
        h_lbl.setObjectName("cardTitle")
        h_top.addWidget(h_lbl)
        h_top.addStretch(1)

        btn_add = QPushButton("选择文件", card1)
        btn_add.setObjectName("primary")
        btn_add.setFixedHeight(34)
        btn_add.setMinimumWidth(120)
        btn_add.clicked.connect(self._choose_input)
        h_top.addWidget(btn_add)

        self._btn_run = QPushButton("开始处理", card1)
        self._btn_run.setObjectName("primary")
        self._btn_run.setFixedHeight(34)
        self._btn_run.setMinimumWidth(120)
        self._btn_run.clicked.connect(self._start_clean)
        h_top.addWidget(self._btn_run)

        self._btn_clear = QPushButton("清除", card1)
        self._btn_clear.setObjectName("primary")
        self._btn_clear.setFixedHeight(34)
        self._btn_clear.setMinimumWidth(120)
        self._btn_clear.clicked.connect(self._clear)
        h_top.addWidget(self._btn_clear)

        c1.addLayout(h_top)

        # 已选文件提示
        self._lbl_file = QLabel("未选择文件", card1)
        self._lbl_file.setObjectName("pageDesc")
        self._lbl_file.setWordWrap(True)
        c1.addWidget(self._lbl_file)

        # 参数行：表头模式 + 分隔符
        param_row = QHBoxLayout()
        param_row.setSpacing(theme.SPACING[16])

        mode_label = QLabel("表头模式:", card1)
        param_row.addWidget(mode_label)
        self._grp_mode = QButtonGroup(self)
        self._radio_single = QRadioButton("单层表头（推荐）", card1)
        self._radio_single.setChecked(True)
        self._radio_double = QRadioButton("双层表头（保留分组行）", card1)
        self._grp_mode.addButton(self._radio_single)
        self._grp_mode.addButton(self._radio_double)
        param_row.addWidget(self._radio_single)
        param_row.addWidget(self._radio_double)

        sep_label = QLabel("分隔符:", card1)
        param_row.addWidget(sep_label)
        self._edit_sep = QLineEdit(card1)
        self._edit_sep.setText("-")
        self._edit_sep.setFixedWidth(48)
        self._edit_sep.setFixedHeight(30)
        param_row.addWidget(self._edit_sep)

        param_row.addStretch(1)
        c1.addLayout(param_row)

        root.addWidget(card1)

        # === 卡片2：结果预览 ===
        card2 = QFrame(self)
        card2.setObjectName("card")
        c2 = QVBoxLayout(card2)
        c2.setContentsMargins(theme.SPACING[16], theme.SPACING[16],
                              theme.SPACING[16], theme.SPACING[16])
        c2.setSpacing(theme.SPACING[12])

        p_head = QHBoxLayout()
        p_head.setSpacing(theme.SPACING[12])
        p_lbl = QLabel("结果预览（前 200 行）", card2)
        p_lbl.setObjectName("cardTitle")
        p_head.addWidget(p_lbl)
        self._lbl_row_count = QLabel("", card2)
        self._lbl_row_count.setObjectName("pageDesc")
        p_head.addWidget(self._lbl_row_count)
        p_head.addStretch(1)
        self._btn_open = QPushButton("打开输出文件", card2)
        self._btn_open.setObjectName("primary")
        self._btn_open.setFixedHeight(34)
        self._btn_open.setMinimumWidth(140)
        self._btn_open.setEnabled(False)
        self._btn_open.clicked.connect(self._open_output)
        p_head.addWidget(self._btn_open)
        c2.addLayout(p_head)

        self._table = _ResultTableView(card2)
        self._table.setObjectName("dataTable")
        self._table.setMinimumHeight(180)
        self._model = QStandardItemModel(self)
        self._table.setModel(self._model)
        c2.addWidget(self._table, stretch=1)

        root.addWidget(card2, stretch=3)

        # === 卡片3：运行日志 ===
        log_card = QFrame(self)
        log_card.setObjectName("card")
        lc = QVBoxLayout(log_card)
        lc.setContentsMargins(theme.SPACING[12], theme.SPACING[12],
                              theme.SPACING[12], theme.SPACING[12])
        lc.setSpacing(6)
        log_title = QLabel("运行日志", log_card)
        log_title.setObjectName("cardTitle")
        lc.addWidget(log_title)
        self._log = QPlainTextEdit(log_card)
        self._log.setObjectName("logView")
        self._log.setReadOnly(True)
        self._log.setMinimumHeight(50)
        lc.addWidget(self._log, stretch=1)
        self._progress = QLabel("", log_card)
        self._progress.setObjectName("pageDesc")
        lc.addWidget(self._progress)

        root.addWidget(log_card, stretch=1)

    # ====== 拖拽 ======

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QDropEvent) -> None:
        for url in event.mimeData().urls():
            fp = url.toLocalFile()
            if fp and fp.lower().endswith((".xlsx", ".xlsm", ".xls")):
                self._set_input(Path(fp))
                break
        event.acceptProposedAction()

    # ====== 文件选择 ======

    def _choose_input(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "选择源 Excel", str(Path.home()), "Excel 文件 (*.xlsx *.xlsm *.xls)")
        if path:
            self._set_input(Path(path))

    def _set_input(self, path: Path) -> None:
        self._input_file = path
        self._lbl_file.setText(f"已选择：{path}")
        self._log_line(f"已选择源文件：{path}")
        if self._output_file is None:
            out = path.parent / f"{path.stem}_清洗.xlsx"
            self._output_file = out
            self._log_line(f"默认输出文件：{out}")

    # ====== 日志 ======

    def _ts(self) -> str:
        return datetime.now().strftime("%H:%M:%S")

    def _log_line(self, msg: str) -> None:
        self._log.appendPlainText(f"[{self._ts()}] {msg}")

    # ====== 清除 ======

    def _clear(self) -> None:
        self._input_file = None
        self._output_file = None
        self._model.clear()
        self._lbl_file.setText("未选择文件")
        self._lbl_row_count.setText("")
        self._log.clear()
        self._progress.setText("")
        self._btn_open.setEnabled(False)
        self._log_line("已清除。")

    # ====== 开始处理 ======

    def _start_clean(self) -> None:
        if self._input_file is None or not self._input_file.exists():
            utils.warn("提示", "请先选择源 Excel 文件。", parent=self)
            return
        if self._output_file is None:
            self._output_file = self._input_file.parent / f"{self._input_file.stem}_清洗.xlsx"

        header_mode = "double" if self._radio_double.isChecked() else "single"
        sep = self._edit_sep.text() or "-"

        self._btn_run.setEnabled(False)
        self._progress.setText("正在处理…")
        self._model.clear()
        self._btn_open.setEnabled(False)
        self._log.clear()
        self._log_line(f"开始清洗（表头模式：{header_mode}，分隔符：{sep!r}）…")

        self._log_queue = Queue()
        self._future = self._executor.submit(
            _logic.process_inventory,
            self._input_file,
            self._output_file,
            header_mode,
            sep,
            self._log_queue.put,
        )

        def _check() -> None:
            while not self._log_queue.empty():
                try:
                    self._log_line(self._log_queue.get_nowait())
                except Exception:
                    break
            if self._future.done():
                try:
                    result = self._future.result()
                    self._on_done(result)
                except Exception as e:  # noqa: BLE001
                    self._on_fail(str(e))
                if self._poll_timer is not None:
                    self._poll_timer.stop()
                    self._poll_timer.deleteLater()
                    self._poll_timer = None

        self._poll_timer = QTimer(self)
        self._poll_timer.setInterval(100)
        self._poll_timer.timeout.connect(_check)
        self._poll_timer.start()

    def _on_done(self, result: _logic.CleanResult) -> None:
        while not self._log_queue.empty():
            try:
                self._log_line(self._log_queue.get_nowait())
            except Exception:
                break
        self._btn_run.setEnabled(True)
        self._progress.setText("")
        self._load_preview(result.output_path, result.header_mode, result.headers)
        self._btn_open.setEnabled(True)
        self._log_line(
            f"✅ 完成：共 {result.header_count} 列表头，{result.data_rows} 行数据。"
        )
        utils.info(
            "处理完成",
            f"输出文件：{result.output_path.name}\n"
            f"表头 {result.header_count} 列，数据 {result.data_rows} 行。",
            parent=self,
        )

    def _on_fail(self, err: str) -> None:
        while not self._log_queue.empty():
            try:
                self._log_line(self._log_queue.get_nowait())
            except Exception:
                break
        self._btn_run.setEnabled(True)
        self._progress.setText("")
        self._log_line(f"❌ 处理失败：{err}")
        utils.error("处理失败", f"清洗时出错：{err}", parent=self)

    # ====== 预览 ======

    def _load_preview(self, path: Path, header_mode: str, headers: list[str]) -> None:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            header_rows = 1 if header_mode == "single" else 2
            if headers:
                col_headers = headers
            else:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
                col_headers = ["" if v is None else str(v) for v in first]
            self._model.setColumnCount(len(col_headers))
            self._model.setHorizontalHeaderLabels(col_headers)
            self._model.setRowCount(0)

            shown = 0
            for row in ws.iter_rows(min_row=header_rows + 1, values_only=True):
                if shown >= self._PREVIEW_MAX_ROWS:
                    break
                items = []
                for v in row:
                    if v is None:
                        items.append("")
                    elif isinstance(v, float) and v.is_integer():
                        items.append(str(int(v)))
                    else:
                        items.append(str(v))
                self._model.appendRow([QStandardItem(x) for x in items])
                shown += 1
            wb.close()
            self._lbl_row_count.setText(f"已显示 {shown} 行（共 {self._model.rowCount()} 行预览）")
        except Exception as e:  # noqa: BLE001
            self._log_line(f"⚠ 预览加载失败：{e}")

    # ====== 打开输出 ======

    def _open_output(self) -> None:
        if self._output_file and self._output_file.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._output_file)))


class InventoryTableCleanerFeature(FeatureModule):
    """存货收发存汇总表清洗 — 功能模块入口。"""

    name = "存货表清洗"
    icon = "fa5s.table"

    def get_widget(self, parent: QWidget | None = None) -> QWidget:
        return InventoryTableCleanerWidget(parent)
