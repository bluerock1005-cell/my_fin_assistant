#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""features/notes_receivable_import/import_logic.py — 应收票据批量导入纯逻辑层。

核心流程（来自产品截图）：
  ① 多个来源 Excel（格式各异）→ 逐个读取 + 定位表头行
  ② 列名关键字匹配 → 映射到模版字段
  ③ 数据清洗/格式转换（日期、金额、票号等）
  ④ 写入固定模板（保留原格式/样式）
  ⑤ 输出匹配日志（哪些列自动匹配成功/失败）

本模块零 UI 依赖，可单独测试或替换。

使用方式：
  - read_and_transform() : 读取+匹配+清洗 → 返回表格可展示数据（不写文件）
  - write_to_template()   : 将数据写入模板副本（保留样式）
  - process_files()       : 完整流程（①~⑤ 一口气，适合一键模式）
"""
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# ===== 模板字段定义 =====

@dataclass
class TemplateField:
    """模板中的一个目标列定义。"""
    col_letter: str           # Excel 列字母 (A, B, C, ...)
    header: str               # 模板表头原文（如 "*签发日"）
    required: bool            # 是否必录（带 * 号）
    aliases: list[str] = field(default_factory=list)


# 必录字段别名映射
_REQUIRED_FIELDS: dict[str, list[str]] = {
    "导入序号": ["导入序号", "序号", "编号", "no", "id", "行号"],
    "票据类型": ["票据类型", "类型", "票种", "汇票类型"],
    "票据号": ["票据号", "票据号码", "号码", "票号", "汇票号码", "发票号码",
               "凭证号", "note_no", "invoice_no"],
    "币别": ["币别", "币种", "货币", "currency"],
    "汇率": ["汇率", "exchange_rate"],
    "签发日": ["签发日", "签发日期", "出票日期", "开票日", "开票日期",
              "issue_date", "出票日"],
    "到期日": ["到期日", "到期日期", "maturity_date", "maturity"],
    "票面金额": ["票面金额", "金额", "面额", "面值", "金额(元)", "amt",
                "amount", "票面"],
    "出票人": ["出票人", "出票单位", "drawer", "issuer"],
    "承兑人": ["承兑人", "承兑银行", "acceptor", "承兑方"],
    "收票日": ["收票日", "收票日期", "收到日", "收到日期"],
    "收款组织": ["收款组织", "收款方组织", "收款公司", "本公司",
                 "我方公司", "本单位"],
    "结算组织": ["结算组织", "结算方组织", "结算公司"],
    "销售组织": ["销售组织", "销售部门组织", "sales_org"],
    "付款单位": ["付款单位", "付款人", "payer", "付款方"],
    "背书明细/导入序号": ["背书明细/导入序号", "背书序号", "endorse_seq"],
}

_OPTIONAL_FIELDS: dict[str, list[str]] = {
    "票面利率(%)": ["票面利率", "利率", "rate", "利率(%)"],
    "承兑协议编号": ["承兑协议编号", "协议编号", "acceptance_agreement_no"],
    "承兑日期": ["承兑日期", "承兑日", "acceptance_date"],
    "期初票据": ["期初票据", "期初", "opening"],
    "带追索权": ["带追索权", "追索权", "recourse"],
    "可撤销": ["可撤销", "撤销", "revocable"],
    "电子票据": ["电子票据", "电子票", "electronic"],
    "销售部门": ["销售部门", "sales_dept"],
    "销售员": ["销售员", "销售人员", "salesman", "业务员"],
    "收款银行网点": ["收款银行网点", "收款开户行", "收款银行"],
    "收款银行账号": ["收款银行账号", "收款账号", "收款账户"],
    "付款银行网点": ["付款银行网点", "付款开户行", "付款银行"],
    "付款银行账号": ["付款银行账号", "付款账号", "付款账户"],
    "付款银行账户名称": ["付款银行账户名称", "付款账户名", "付款户名"],
    "应收票据明细/背书日期": ["背书日期", "endorse_date"],
    "应收票据明细/背书人（前手）": ["背书人", "前手", "endorser"],
}


def _build_template_fields() -> list[TemplateField]:
    """按模板列顺序 (A → AG) 构建完整字段列表。"""
    _col_headers = [
        ("*导入序号", True),
        ("*票据类型", True),
        ("*票据号", True),
        ("*币别", True),
        ("*汇率", True),
        ("*签发日", True),
        ("*到期日", True),
        ("*票面金额", True),
        ("票面利率(%)", False),
        ("*出票人", True),
        ("*承兑人", True),
        ("承兑协议编号", False),
        ("承兑日期", False),
        ("期初票据", False),
        ("带追索权", False),
        ("可撤销", False),
        ("电子票据", False),
        ("*收票日", True),
        ("*收款组织", True),
        ("*结算组织", True),
        ("*付款单位", True),
        ("*付款单位多资料值", True),
        ("*销售组织", True),
        ("销售部门", False),
        ("销售员", False),
        ("收款银行网点", False),
        ("收款银行账号", False),
        ("付款银行网点", False),
        ("付款银行账号", False),
        ("付款银行账户名称", False),
        ("*背书明细/导入序号", True),
        ("应收票据明细/背书日期", False),
        ("应收票据明细/背书人（前手）", False),
    ]
    all_aliases = {**_REQUIRED_FIELDS, **_OPTIONAL_FIELDS}
    fields = []
    for idx, (header, required) in enumerate(_col_headers, start=1):
        name = header.lstrip("*")
        fields.append(TemplateField(
            col_letter=get_column_letter(idx),
            header=header,
            required=required,
            aliases=all_aliases.get(name, [name]),
        ))
    return fields


TEMPLATE_FIELDS = _build_template_fields()

# 公开属性供 UI 层获取表头列表
TEMPLATE_HEADERS = [tf.header for tf in TEMPLATE_FIELDS]
TEMPLATE_HEADERS_DISPLAY = [tf.header.lstrip("*") for tf in TEMPLATE_FIELDS]
REQUIRED_FIELD_SET = {tf.header for tf in TEMPLATE_FIELDS if tf.required}

# ===== 固定字段（不参与映射配置，由系统按规则自动填充）=====
# 这些字段在「配置映射」对话框中不会出现，始终按规则填充：
#   "auto_seq"  → 按记录顺序自动编号 (1, 2, 3, ...)
#   "blank"     → 留空
#   其他值      → 固定常量值（可为字符串或数字，如 汇率=1.0）
FIXED_FIELDS: dict[str, Any] = {
    "*导入序号": "auto_seq",
    "*票据类型": "银行承兑汇票",
    "*币别": "人民币",
    "*背书明细/导入序号": "blank",
    "*汇率": 1.0,
    # 付款单位固定默认「客户」，不进入配置映射对话框、不参与来源列匹配
    "*付款单位": "客户",
}
FIXED_FIELD_HEADERS = set(FIXED_FIELDS.keys())

# ===== 页面输入字段（不进映射匹配，取值来自主界面用户填写）=====
# 这些字段在「配置映射」对话框中显示为只读（实际值 = 页面「收款组织」文本框），
# 不参与来源列匹配，也不计入「缺少必录」提示；其中「结算组织」「销售组织」
# 恒等于「收款组织」（销售组织按用户规定必填且 = 收款组织）。
PAGE_INPUT_FIELDS: set[str] = {"收款组织", "结算组织", "销售组织"}
# 对应模板表头（带 *，与 TEMPLATE_FIELDS 中的 header 一致）
PAGE_INPUT_HEADERS: set[str] = {f"*{n}" for n in PAGE_INPUT_FIELDS}

# ===== 默认不匹配字段 =====
# 这些字段虽是必录，但来源表头写法差异太大、自动匹配常常出错，
# 因此 match_columns 自动匹配阶段跳过它们，默认「不匹配/固定值」，
# 由用户在「字段映射配置」中手动选择来源列。
# 注意：仍可被用户的手工映射（manual_map）命中，保存后持久化生效。
NO_AUTO_MATCH_HEADERS: set[str] = {"*收票日", "*付款单位多资料值"}

# 用于必录缺值红标 / 缺失必录判断时排除固定字段（它们由系统保证填充）。
# 例：*背书明细/导入序号 固定留空，不应被标红或列为缺录。
REQUIRED_FOR_CHECK = REQUIRED_FIELD_SET - FIXED_FIELD_HEADERS


# ===== 匹配结果数据结构 =====

@dataclass
class ColumnMatch:
    source_header: str
    target_field: str
    target_col: str
    confidence: float
    is_auto: bool


@dataclass
class FileResult:
    file_name: str
    file_path: Path
    total_rows: int = 0
    matched_columns: list[ColumnMatch] = field(default_factory=list)
    unmatched_source: list[str] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    cleaned_data: list[dict[str, Any]] = field(default_factory=list)
    source_headers: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    """完整批量导入结果。"""
    files: list[FileResult] = field(default_factory=list)
    output_path: Path | None = None
    total_imported: int = 0
    log_lines: list[str] = field(default_factory=list)
    # 新增：表格展示用的扁平数据（list[dict], key=模板字段header）
    table_data: list[dict[str, Any]] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return any(f.errors for f in self.files)


# ===== ① 读取来源文件 + 定位表头 =====


def _find_header_row(ws, max_scan: int = 20) -> int | None:
    """在工作表中定位表头行。"""
    best_row = None
    best_non_empty = 0
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]
        non_empty = sum(1 for v in row_values if v is not None and str(v).strip())
        if non_empty > best_non_empty:
            text = " ".join(str(v) for v in row_values if v is not None)
            if re.search(r"[\u4e00-\u9fff]|序号|编号|日期|金额|票据|类型|币别", text):
                best_non_empty = non_empty
                best_row = row_idx
    return best_row


def read_source_file(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    """读取来源 Excel（.xlsx/.xls），返回 (header_list, data_rows)。

    自动探测表头行位置；.xls 格式尝试用 xlrd 回退。
    """
    ext = file_path.suffix.lower()
    if ext == ".xls":
        return _read_xls(file_path)
    # 默认用 openpyxl (.xlsx/.xlsm)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    header_row_idx = _find_header_row(ws) or 1
    headers, rows = _extract_rows(ws, header_row_idx)
    wb.close()
    return headers, rows


def _read_xls(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    """读取旧版 .xls 文件（优先尝试 openpyxl，失败则 xlrd 回退）。"""
    # 先尝试 openpyxl（某些 .xls 其实可以用）
    try:
        return read_source_file_with_openpyxl(file_path)
    except Exception:
        pass
    # 尝试 xlrd
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path, formatting_info=False)
        ws = wb.sheet_by_index(0)
        # 找表头行
        header_row_idx = 0
        best_non_empty = 0
        for r in range(min(20, ws.nrows)):
            vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            non_empty = sum(1 for v in vals if v and str(v).strip())
            text = " ".join(str(v) for v in vals if v)
            if non_empty > best_non_empty and re.search(
                r"[\u4e00-\u9fff]|序号|编号|日期|金额|票据|类型|币别", text
            ):
                best_non_empty = non_empty
                header_row_idx = r
        headers: list[str] = []
        rows: list[list[Any]] = []
        for r in range(ws.nrows):
            vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if r == header_row_idx:
                headers = [_coerce_str(v) for v in vals]
            elif r > header_row_idx:
                if any(v and str(v).strip() for v in vals):
                    rows.append(vals)
        return headers, rows
    except ImportError:
        raise ValueError(
            f"无法读取 .xls 文件: {file_path.name}\n请安装 xlrd: pip install xlrd"
        )
    except Exception as e:
        raise ValueError(f"无法读取 .xls 文件: {e}")


def _coerce_str(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # xlrd 可能返回 float 类型的数字字符串如 123.0 → "123"
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return s


def read_source_file_with_openpyxl(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    """纯 openpyxl 读取路径（供 _read_xls 内部回退调用）。"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    hri = _find_header_row(ws) or 1
    headers, rows = _extract_rows(ws, hri)
    wb.close()
    return headers, rows


def _extract_rows(ws, header_row_idx: int) -> tuple[list[str], list[list[Any]]]:
    """从已打开的工作表中提取表头和数据行。"""
    headers: list[str] = []
    rows: list[list[Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [v for v in row]
        if r_idx == header_row_idx:
            headers = [str(v).strip() if v is not None else "" for v in values]
        elif r_idx > header_row_idx:
            if any(v is not None and str(v).strip() for v in values):
                rows.append(values)
    return headers, rows


# ===== ② 列名关键字匹配 =====


def _score_match(source_name: str, target_aliases: list[str]) -> float:
    """计算来源表头与目标字段别名的最佳匹配分数（0~1）。"""
    s = source_name.strip().lower().replace(" ", "").replace("（", "(").replace("）", ")")
    s_norm = re.sub(r"[()（）\*/]", "", s)
    best = 0.0
    for alias in target_aliases:
        a = alias.strip().lower().replace(" ", "")
        a_norm = re.sub(r"[()（）\*/]", "", a)
        if s == a or s_norm == a_norm:
            return 1.0
        if s in a or a in s or s_norm in a_norm or a_norm in s_norm:
            best = max(best, 0.85)
        else:
            common = sum(1 for ch in s_norm if ch in a_norm)
            ratio = common / max(len(s_norm), len(a_norm), 1)
            if ratio >= 0.6:
                best = max(best, 0.6 + ratio * 0.2)
            elif s_norm[:1] == a_norm[:1] and abs(len(s_norm) - len(a_norm)) <= 2:
                best = max(best, 0.5)
    return best


def match_columns(
    source_headers: list[str],
    manual_map: dict[str, str] | None = None,
    excluded_targets: set[str] | None = None,
) -> tuple[list[ColumnMatch], list[str], list[str]]:
    """将来源表头映射到模板字段。

    Returns:
        (matched_columns, unmatched_sources, missing_required)
    """
    manual_map = manual_map or {}
    matched: list[ColumnMatch] = []
    used_targets: set[str] = set()
    used_sources: set[str] = set()

    # 手动指定优先（固定字段由系统填充，忽略用户的手工映射）
    for src, tgt_name in manual_map.items():
        tgt_field = next(
            (f for f in TEMPLATE_FIELDS
             if f.header.lstrip("*") == tgt_name or f.header == tgt_name), None
        )
        if (tgt_field and tgt_field.header not in FIXED_FIELD_HEADERS
                and tgt_field.header not in PAGE_INPUT_HEADERS
                and src not in used_sources):
            matched.append(ColumnMatch(src, tgt_field.header, tgt_field.col_letter, 1.0, False))
            used_targets.add(tgt_field.header)
            used_sources.add(src)

    # 自动匹配剩余：仅对「必录字段」自动匹配。
    # 非必录字段默认留空（"非必填项默认空白"），只有用户显式在映射对话框
    # 中选了来源列（写入 manual_map）才会被匹配，避免无谓地把来源列
    # 自动绑到可选字段上。
    candidates: list[tuple[float, ColumnMatch]] = []
    for src in source_headers:
        if src in used_sources or not src.strip():
            continue
        for tf in TEMPLATE_FIELDS:
            if tf.header in used_targets:
                continue
            if tf.header in FIXED_FIELD_HEADERS:
                continue
            if tf.header in PAGE_INPUT_HEADERS:
                continue
            if tf.header in NO_AUTO_MATCH_HEADERS:
                continue
            if excluded_targets and tf.header in excluded_targets:
                continue
            if not tf.required:
                continue
            score = _score_match(src, tf.aliases)
            if score >= 0.5:
                cm = ColumnMatch(src, tf.header, tf.col_letter, score, True)
                candidates.append((-score, cm))

    candidates.sort(key=lambda x: x[0])
    for _, cm in candidates:
        if cm.target_field not in used_targets and cm.source_header not in used_sources:
            matched.append(cm)
            used_targets.add(cm.target_field)
            used_sources.add(cm.source_header)

    unmatched = [h for h in source_headers if h.strip() and h not in used_sources]
    missing_req = [tf.header for tf in TEMPLATE_FIELDS
                   if tf.required and tf.header not in used_targets
                   and tf.header not in FIXED_FIELD_HEADERS
                   and tf.header not in PAGE_INPUT_HEADERS]

    return matched, unmatched, missing_req


# ===== ③ 数据清洗 / 格式转换 =====


def _to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("，", "")
    s = re.sub(r"[¥￥元$,\s]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_date(val: Any) -> date | datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
                "%Y年%m月%d日", "%Y%m%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_bool(val: Any) -> bool | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("true", "是", "yes", "1", "y", "√"):
        return True
    if s in ("false", "否", "no", "0", "n", "×", "x"):
        return False
    return None


def clean_value(val: Any, target_field: str) -> Any:
    """根据目标字段类型清洗转换值。返回 str/float/date/bool/None。"""
    name = target_field.lstrip("*")

    if name in ("签发日", "到期日", "收票日", "承兑日期", "应收票据明细/背书日期"):
        d = _to_date(val)
        return d if d is not None else (_to_str(val) if val is not None else None)

    if name in ("票面金额", "汇率", "票面利率(%)"):
        return _to_float(val)

    if name in ("期初票据", "带追索权", "可撤销", "电子票据"):
        return _to_bool(val)

    if name == "导入序号":
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None

    return _to_str(val)


def value_to_display(val: Any) -> str:
    """将任意清洗后的值转为表格展示字符串。"""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "是" if val else "否"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        # 整数浮点不显示小数点
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


# ===== 核心中间函数：读取+变换（不写文件）=====


def read_and_transform(
    file_paths: list[Path],
    defaults: dict[str, Any] | None = None,
    manual_maps: dict[str, dict[str, str]] | None = None,
    excluded_maps: dict[str, list[str]] | None = None,
) -> ImportResult:
    """读取多个来源文件 → 匹配列 → 清洗数据 → 返回结果（含 table_data 供 UI 表格展示）。

    此函数不做文件写入，仅负责「读→匹配→洗」三步。
    UI 层拿到 result.table_data 后可展示到 QTableView，用户确认后再调 write_to_template 导出。
    """
    result = ImportResult()
    defaults = defaults or {}
    manual_maps = manual_maps or {}
    seq = 0
    all_cleaned: list[dict[str, Any]] = []

    for fp in file_paths:
        fr = FileResult(file_name=fp.name, file_path=fp)
        result.files.append(fr)

        try:
            headers, raw_rows = read_source_file(fp)
        except Exception as e:
            fr.errors.append(f"文件读取失败: {e}")
            result.log_lines.append(f"[错误] {fp.name}: 文件读取失败 - {e}")
            continue

        if not headers or not any(h.strip() for h in headers):
            fr.errors.append("未能识别表头行")
            result.log_lines.append(f"[警告] {fp.name}: 未能识别表头行，已跳过")
            continue

        fr.source_headers = headers

        fmap = manual_maps.get(fp.name, {}) if manual_maps else {}
        excl = excluded_maps.get(fp.name) if excluded_maps else None
        matched, unmatched, missing = match_columns(
            headers, fmap if fmap else None, set(excl) if excl else None)
        fr.matched_columns = matched
        fr.unmatched_source = unmatched
        fr.missing_required = missing

        # 日志
        result.log_lines.append(f"\n{'='*60}")
        result.log_lines.append(f"[文件] {fp.name}")
        nh = len([h for h in headers if h.strip()])
        result.log_lines.append(f"  来源列数: {nh}, 数据行数: {len(raw_rows)}")
        if matched:
            result.log_lines.append(f"  ✓ 已匹配 {len(matched)} 列:")
            for m in sorted(matched, key=lambda x: x.confidence, reverse=True):
                flag = "✓" if m.confidence >= 0.8 else "~"
                result.log_lines.append(
                    f"    {flag} [{m.target_col}] {m.target_field} ← "
                    f"\"{m.source_header}\" ({m.confidence:.0%})")
        if unmatched:
            result.log_lines.append(f'  ✗ 未匹配 {len(unmatched)} 列: {", ".join(unmatched)}')
        if missing:
            result.log_lines.append(f'  ⚠ 缺少必录: {", ".join(missing)}')

        # 清洗
        src_to_target: dict[int, str] = {}
        for m in matched:
            try:
                src_idx = headers.index(m.source_header)
                src_to_target[src_idx] = m.target_field
            except ValueError:
                pass

        for raw_row in raw_rows:
            seq += 1
            record: dict[str, Any] = {"*导入序号": seq}
            for src_idx, target_field in src_to_target.items():
                if src_idx < len(raw_row):
                    record[target_field] = clean_value(raw_row[src_idx], target_field)
            # 默认值补充必录空字段
            for tf in TEMPLATE_FIELDS:
                if tf.required:
                    if tf.header not in record or record[tf.header] is None:
                        if tf.header in defaults:
                            record[tf.header] = defaults[tf.header]
                        elif tf.header == "*导入序号":
                            record[tf.header] = seq

            # 页面输入字段：收款组织 = 页面填写值；
            # 结算组织 = 收款组织；销售组织 = 收款组织（用户规定必填且相等）。
            # 三者均不来自来源列（已在 match_columns 中排除匹配），
            # 使用带 * 的表头键，保证 write_to_template 能正确落列。
            recv_header = "*收款组织"
            settle_header = "*结算组织"
            sales_header = "*销售组织"
            _recv = record.get(recv_header)
            if _recv in (None, ""):
                _recv = defaults.get(recv_header) or defaults.get("收款组织")
            if _recv not in (None, ""):
                record[recv_header] = _recv
                record[settle_header] = _recv
                record[sales_header] = _recv

            # 固定字段：始终按规则填充，优先级最高（覆盖来源列/默认值）
            # 注：*付款单位 已在 FIXED_FIELDS 中固定为「客户」
            for hdr, rule in FIXED_FIELDS.items():
                if rule == "auto_seq":
                    record[hdr] = seq
                elif rule == "blank":
                    record[hdr] = ""
                else:
                    record[hdr] = rule

            all_cleaned.append(record)

        fr.total_rows = len(raw_rows)
        fr.cleaned_data = all_cleaned[-len(raw_rows):] if raw_rows else []

    result.table_data = all_cleaned
    if not all_cleaned:
        result.log_lines.append("\n[跳过] 无有效数据")
    else:
        result.log_lines.append(f"\n[就绪] 共 {len(all_cleaned)} 条记录待导出")

    return result


# ===== ④ 写入固定模板 =====

TEMPLATE_FILENAME = "应收票据导入模版.xlsx"


def get_template_path(data_dir: Path | None = None) -> Path:
    """获取内置模板路径。"""
    if data_dir and (data_dir / TEMPLATE_FILENAME).exists():
        return data_dir / TEMPLATE_FILENAME
    fallback = Path(__file__).resolve().parent.parent.parent / TEMPLATE_FILENAME
    if fallback.exists():
        return fallback
    raise FileNotFoundError(f"找不到模板文件: {TEMPLATE_FILENAME}")


def get_template_headers(template_path: Path | None = None,
                         data_dir: Path | None = None) -> list[str]:
    """读取模板文件的表头行（用于动态切换模板时刷新 UI 表格列）。"""
    if template_path is None:
        template_path = get_template_path(data_dir)
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "") for cell in next(ws.iter_rows(max_row=1))]
    wb.close()
    return headers


def write_to_template(
    table_data: list[dict[str, Any]],
    output_path: Path,
    template_path: Path | None = None,
    data_dir: Path | None = None,
) -> int:
    """将表格数据写入模板副本，保留原格式/样式。

    Args:
        table_data: UI 表格中的数据（每项 key=模板字段名, value=任意类型）。
                     与 result.table_data 结构一致。
        output_path: 输出文件路径（直接覆盖，不追加不合并）。
        template_path: 模板文件路径。
        data_dir: 数据目录（查找内置模板）。

    Returns:
        写入的行数。
    """
    if template_path is None:
        template_path = get_template_path(data_dir)

    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    start_row = 2  # 表头在第 1 行

    # 清除模板已有示例数据
    if ws.max_row >= 2:
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c, value=None)

    # 字段→列号映射
    field_to_col: dict[str, int] = {
        tf.header: openpyxl.utils.column_index_from_string(tf.col_letter)
        for tf in TEMPLATE_FIELDS
    }

    for row_idx, record in enumerate(table_data, start=start_row):
        for field_name, value in record.items():
            col_idx = field_to_col.get(field_name)
            if col_idx is not None and value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    wb.close()
    return len(table_data)


# ===== ⑤ 完整流程（一键模式，兼容旧接口）=====


def process_files(
    file_paths: list[Path],
    output_path: Path,
    template_path: Path | None = None,
    data_dir: Path | None = None,
    manual_maps: dict[str, dict[str, str]] | None = None,
    defaults: dict[str, Any] | None = None,
    excluded_maps: dict[str, list[str]] | None = None,
) -> ImportResult:
    """完整流程：读取→匹配→清洗→写入模板→返回结果。"""
    result = read_and_transform(file_paths, defaults, manual_maps, excluded_maps)

    if result.table_data:
        try:
            written = write_to_template(result.table_data, output_path, template_path, data_dir)
            result.total_imported = written
            result.output_path = output_path
            result.log_lines.append(f"\n{'='*60}")
            result.log_lines.append(f"[完成] 共导入 {written} 条记录 → {output_path}")
        except Exception as e:
            result.log_lines.append(f"\n[错误] 写入模板失败: {e}")
            for fr in result.files:
                if not fr.errors:
                    fr.errors.append(f"写入失败: {e}")
    else:
        result.log_lines.append("\n[跳过] 无有效数据可导入")

    return result
