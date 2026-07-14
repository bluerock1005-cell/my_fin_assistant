#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""存货表清洗：read_accounting_periods 单测（无头，不依赖 Excel COM）。

验证：源文件 B 列（会计期间）读取 → 去重 → 排序 → 返回零填充字符串列表。
"""
from __future__ import annotations

import warnings

from pathlib import Path

import openpyxl
import pytest

from features.inventory_table_cleaner import inventory_cleaner_logic as logic

# 抑制 openpyxl 在无默认样式文件上的 UserWarning
warnings.filterwarnings("ignore", category=UserWarning)

_SAMPLE = Path("测试文档/存货收发存汇总表测试.xlsx").resolve()


def test_sample_file_periods():
    """样例文件 B 列应为 01~06，去重排序后返回零填充字符串。"""
    periods = logic.read_accounting_periods(_SAMPLE)
    assert periods == ["01", "02", "03", "04", "05", "06"]
    # 全部为零填充双字符
    assert all(len(p) == 2 and p.isdigit() for p in periods)


def test_single_digit_normalized(tmp_path: Path):
    """B 列为单数字（"1".."4"）应归一化为零填充 "01".."04"。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "会计年度", "会计期间"
    for i in range(1, 5):
        ws.cell(i + 1, 1, "2026")
        ws.cell(i + 1, 2, str(i))  # "1".."4"
    tmp = tmp_path / "inv_periods.xlsx"
    wb.save(tmp)
    assert logic.read_accounting_periods(tmp) == ["01", "02", "03", "04"]


def test_merged_title_row_excluded(tmp_path: Path):
    """整行合并的标题（B 列含长文本）不应被当成会计期间。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 第 1 行整行合并（标题）
    ws["A1"] = "核算体系：xxx  会计期间：2026年01期 - 2026年06期"
    ws.merge_cells("A1:K1")
    ws["A2"], ws["B2"] = "会计年度", "会计期间"
    for i in range(1, 4):
        ws.cell(i + 2, 1, "2026")
        ws.cell(i + 2, 2, f"0{i}")  # "01".."03"
    tmp = tmp_path / "inv_merged.xlsx"
    wb.save(tmp)
    assert logic.read_accounting_periods(tmp) == ["01", "02", "03"]


def test_missing_file_raises():
    with pytest.raises(FileNotFoundError):
        logic.read_accounting_periods("不存在的文件.xlsx")


def test_find_open_cols():
    headers = [
        "会计年度", "会计期间", "存货类别",
        "期初结存-数量", "期初结存-单价", "期初结存-金额",
        "本期收入-数量", "本期结存-数量",
    ]
    assert logic._find_open_cols(headers, "-") == [4, 5, 6]


def test_find_close_cols_prefers_final():
    """表头同时有 期末结存 / 本期结存 时优先用 期末结存。"""
    headers = [
        "会计年度", "会计期间",
        "期初结存-数量", "期初结存-单价", "期初结存-金额",
        "期末结存-数量", "期末结存-单价", "期末结存-金额",
        "本期结存-数量", "本期结存-单价", "本期结存-金额",
    ]
    assert logic._find_close_cols(headers, "-") == [6, 7, 8]


def test_find_close_cols_fallback_to_current():
    """只有 本期结存 时也能找到。"""
    headers = [
        "会计年度", "会计期间",
        "期初结存-数量", "本期结存-数量", "本期结存-单价", "本期结存-金额",
    ]
    assert logic._find_close_cols(headers, "-") == [4, 5, 6]


def _headers_with_close_at(target: list[str], ncols: int = 57) -> list[str]:
    """构造 ncols 列的表头，把 BC/BD/BE 设为 target 三列。"""
    h = [""] * ncols
    h[1] = "会计期间"
    h[54], h[55], h[56] = target  # BC, BD, BE（0-based 索引）
    return h


def test_resolve_close_cols_explicit_bc_bd_be():
    """标准表头：BC/BD/BE 为本期结存数量/单价/金额，应显式锁定 55/56/57。"""
    headers = _headers_with_close_at(
        ["本期结存-数量", "本期结存-单价", "本期结存-金额"])
    assert logic._resolve_close_cols(headers, "-") == [55, 56, 57]


def test_resolve_close_cols_explicit_when_dynamic_would_fail():
    """多期间、表头带期号时动态匹配失效，但显式 BC/BD/BE 仍能命中。"""
    # BC/BD/BE 变成带期号的「本期结存-06-数量」，且表里没有不含期号的
    # 「本期结存-数量」——此时 _find_close_cols 会返回 None。
    headers = _headers_with_close_at(
        ["本期结存-06-数量", "本期结存-06-单价", "本期结存-06-金额"])
    assert logic._find_close_cols(headers, "-") is None
    # 显式路径应忽略表头文本、直接锁定 BC/BD/BE
    assert logic._resolve_close_cols(headers, "-") == [55, 56, 57]


def test_resolve_close_cols_fallback_when_too_narrow():
    """表头不足 57 列（BC/BD/BE 不存在）时回退动态查找。"""
    headers = ["会计年度", "会计期间",
               "期初结存-数量", "本期结存-数量", "本期结存-单价", "本期结存-金额"]
    assert logic._resolve_close_cols(headers, "-") == [4, 5, 6]


def test_resolve_close_cols_fallback_when_not_jiecun():
    """BC/BD/BE 表头不含「结存」时回退动态查找。"""
    headers = _headers_with_close_at(["其他列-A", "其他列-B", "其他列-C"])
    assert logic._resolve_close_cols(headers, "-") is None


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
