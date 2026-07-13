#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""notes_receivable_import 读取健壮性回归测试。

覆盖一个真实踩坑场景：
  某些系统导出的 .xlsx，其 sheet 的 <dimension> 属性被错误地写成
  "A1"（而实际表格有 N 行 M 列）。openpyxl 的 read_only=True 模式
  会信任该属性，把整表误判为 1 行 1 列，导致表头/数据全部读空、
  表现为「导入后无数据、无法配置映射」。

本测试手工构造这样一个「dimension 损坏」的 xlsx，断言 read_source_file
能经回退路径（非 read_only 完整加载）恢复真实尺寸。
"""
import os
import sys
import zipfile
import shutil
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from pathlib import Path

from features.notes_receivable_import import import_logic as L

_TMP = Path(tempfile.gettempdir()) / "notes_import_dim_test"
_HEADERS = ["票据（包）号", "子票区间", "票据（包）金额(元)",
            "出票日期", "汇票到期日", "能否转让", "出票人全称",
            "承兑人全称", "持票人全称", "申请人", "应答日期",
            "票据状态", "流通标志"]
_ROWS = [
    ["A001", "1-5", 100.0, "2026-01-01", "2026-06-01", "可", "甲", "乙", "丙", "丁", "2026-01-02", "正常", "流通"],
    ["A002", "6-9", 200.5, "2026-02-01", "2026-07-01", "否", "戊", "己", "庚", "辛", "2026-02-02", "正常", "流通"],
    ["A003", "1-3", 300.0, "2026-03-01", "2026-08-01", "可", "壬", "癸", "子", "丑", "2026-03-02", "挂失", "冻结"],
]


def _corrupt_dimension(src: Path, dst: Path) -> None:
    """复制 xlsx 并把 sheet1.xml 的 <dimension> 改成 "A1"。"""
    with zipfile.ZipFile(src, "r") as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}
    sheet = [n for n in names if n.startswith("xl/worksheets/sheet")][0]
    xml = data[sheet].decode("utf-8", "replace")
    import re
    xml = re.sub(r'<dimension ref="[^"]*"/>', '<dimension ref="A1"/>', xml)
    data[sheet] = xml.encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, b in data.items():
            zout.writestr(n, b)


def _make_corrupted() -> Path:
    _TMP.mkdir(parents=True, exist_ok=True)
    good = _TMP / "_good.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for r in _ROWS:
        ws.append(r)
    wb.save(good)
    wb.close()
    bad = _TMP / "broken_dimension.xlsx"
    _corrupt_dimension(good, bad)
    return bad


def test_read_recovers_broken_dimension():
    fp = _make_corrupted()
    # 确认 read_only 模式确实会被误导（复现 bug 前提）
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    assert ws.max_row <= 1 and ws.max_column <= 1, "前提不成立：read_only 未误判为 1x1"
    wb.close()

    # 正式断言：read_source_file 必须恢复真实尺寸
    headers, rows = L.read_source_file(fp)
    assert len(headers) == len(_HEADERS), (
        f"表头列数应为 {len(_HEADERS)}，实际 {len(headers)}")
    assert headers == _HEADERS, f"表头内容不匹配：{headers}"
    assert len(rows) == len(_ROWS), (
        f"数据行数应为 {len(_ROWS)}，实际 {len(rows)}")
    # 首行数据应正确
    assert rows[0][0] == "A001" and rows[0][2] == 100.0, "首行数据解析异常"
    print("  ✓ read_source_file 从 dimension='A1' 损坏文件恢复 13列/3行 成功")


if __name__ == "__main__":
    try:
        test_read_recovers_broken_dimension()
    finally:
        if _TMP.exists():
            shutil.rmtree(_TMP, ignore_errors=True)
    print("\nALL TESTS PASSED")
